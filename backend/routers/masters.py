"""Master data CRUD — generic factory for all 14 entity types."""

from io import BytesIO
from datetime import datetime

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from openpyxl import Workbook, load_workbook

from db import get_db
from models import (
    Vendor, CostCenter, GLAccount, TaxCode, CompanyCode, PlantCode,
    ApprovalRule, BusinessRuleConfig, FreightRateCard, ServiceRateCard,
    AgreementMaster, InvoiceCategoryConfig, ApprovalSequenceMaster, new_id,
)

router = APIRouter(prefix="/api/masters", tags=["masters"])

# Field mapping: camelCase body keys -> snake_case model attrs
# Only needed for non-trivial mappings; simple ones (name, code, etc.) handled by lowercasing
FIELD_MAP = {
    "vendorNumber": "vendor_number", "taxId": "tax_id", "paymentTerms": "payment_terms",
    "bankAccount": "bank_account", "branchCode": "branch_code", "isActive": "is_active",
    "companyCode": "company_code", "accountNumber": "account_number",
    "minAmount": "min_amount", "maxAmount": "max_amount",
    "requiredApprovers": "required_approvers", "approverIds": "approver_ids",
    "slaHours": "sla_hours", "containerType": "container_type",
    "vendorId": "vendor_id", "vendorName": "vendor_name",
    "agreementNumber": "agreement_number", "startDate": "start_date",
    "endDate": "end_date", "requiredDocs": "required_docs",
    "extractionTemplateId": "extraction_template_id",
    "authChainId": "auth_chain_id", "glAccount": "gl_account",
    "invoiceType": "invoice_type",
}


def _crud(model_cls, prefix: str, id_prefix: str):
    """Register GET (list), POST (create), PUT (update), DELETE for a model."""

    @router.get(f"/{prefix}")
    def list_all(db: Session = Depends(get_db)):
        return [r.to_dict() for r in db.query(model_cls).all()]

    @router.post(f"/{prefix}")
    def create(body: dict, db: Session = Depends(get_db)):
        obj = model_cls(id=new_id(id_prefix))
        _apply_body(obj, body)
        db.add(obj)
        db.commit()
        return obj.to_dict()

    @router.put(f"/{prefix}/{{item_id}}")
    def update(item_id: str, body: dict, db: Session = Depends(get_db)):
        obj = db.query(model_cls).filter(model_cls.id == item_id).first()
        if not obj:
            raise HTTPException(404, f"{model_cls.__name__} not found")
        _apply_body(obj, body)
        db.commit()
        return obj.to_dict()

    @router.delete(f"/{prefix}/{{item_id}}")
    def delete(item_id: str, db: Session = Depends(get_db)):
        obj = db.query(model_cls).filter(model_cls.id == item_id).first()
        if not obj:
            raise HTTPException(404, f"{model_cls.__name__} not found")
        db.delete(obj)
        db.commit()
        return {"success": True}

    # Rename functions to avoid FastAPI route conflicts
    list_all.__name__ = f"list_{prefix}"
    create.__name__ = f"create_{prefix}"
    update.__name__ = f"update_{prefix}"
    delete.__name__ = f"delete_{prefix}"


def _apply_body(obj, body: dict):
    """Apply camelCase body dict to snake_case model attributes."""
    for key, val in body.items():
        if key == "id":
            continue
        attr = FIELD_MAP.get(key, key)
        if hasattr(obj, attr):
            setattr(obj, attr, val)


# Register all master data CRUD routes
_crud(Vendor, "vendors", "VND-")
_crud(CostCenter, "cost-centers", "CC-")
_crud(GLAccount, "gl-accounts", "GL-")
_crud(TaxCode, "tax-codes", "TX-")
_crud(CompanyCode, "company-codes", "COMP-")
_crud(PlantCode, "plant-codes", "PLT-")
_crud(ApprovalRule, "approval-rules", "AR-")
_crud(BusinessRuleConfig, "business-rule-configs", "BR-")
_crud(FreightRateCard, "freight-rate-cards", "FRC-")
_crud(ServiceRateCard, "service-rate-cards", "SRC-")
_crud(AgreementMaster, "agreement-masters", "AGR-")
_crud(InvoiceCategoryConfig, "invoice-category-configs", "ICC-")
_crud(ApprovalSequenceMaster, "approval-sequences", "ASM-")


# ---------------------------------------------------------------------------
# Master type registry for upload/download
# ---------------------------------------------------------------------------
MASTER_TYPES = {
    "vendors": Vendor,
    "freight-rate-cards": FreightRateCard,
    "service-rate-cards": ServiceRateCard,
    "approval-sequences": ApprovalSequenceMaster,
}


def _cell_val(v):
    """Convert cell value to JSON-safe type."""
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.isoformat()
    return v


# ---------------------------------------------------------------------------
# Upload — POST /api/masters/{master_type}/upload
# ---------------------------------------------------------------------------
@router.post("/{master_type}/upload")
async def upload_master(master_type: str, file: UploadFile = File(...), db: Session = Depends(get_db)):
    if master_type not in MASTER_TYPES:
        raise HTTPException(400, f"Invalid master_type: {master_type}. Must be one of: {list(MASTER_TYPES.keys())}")

    data = await file.read()
    wb = load_workbook(BytesIO(data), read_only=True, data_only=True)

    if master_type == "vendors":
        return _upload_vendors(wb, db)
    elif master_type == "freight-rate-cards":
        return _upload_freight(wb, db)
    elif master_type == "service-rate-cards":
        return _upload_service(wb, db)
    elif master_type == "approval-sequences":
        return _upload_approval(wb, db)


def _upload_vendors(wb, db: Session):
    """Parse vendor Excel (Sheet1 = data, row 1 = headers)."""
    ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"inserted": 0, "updated": 0, "skipped": 0, "errors": []}

    headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(rows[0])]
    inserted, updated, errors = 0, 0, []

    for row_num, row in enumerate(rows[1:], start=2):
        try:
            raw = {h: _cell_val(v) for h, v in zip(headers, row)}
            vn = str(raw.get("Vendor Number", "")).strip()
            if not vn:
                errors.append({"row": row_num, "message": "Missing Vendor Number"})
                continue

            existing = db.query(Vendor).filter(Vendor.vendor_number == vn).first()
            obj = existing or Vendor(id=new_id("VND-"), vendor_number=vn)

            obj.name = str(raw.get("Name1", "")).strip()
            obj.tax_id = str(raw.get("Tax Number 1", "")).strip()
            obj.address = str(raw.get("Street", "")).strip()
            obj.city = str(raw.get("City", "")).strip()
            obj.country = str(raw.get("Country", "")).strip() or "Australia"
            obj.payment_terms = str(raw.get("Paymt Term Company", "") or raw.get("Payment Term", "")).strip()
            obj.bank_key = str(raw.get("Bank Key", "")).strip()
            obj.bank_account_number = str(raw.get("Bank Account Number", "")).strip()
            obj.email = str(raw.get("E-mail Address", "")).strip()
            obj.currency = str(raw.get("Order Currency", "")).strip() or "AUD"
            obj.company_code = str(raw.get("Company Code", "")).strip()
            obj.is_active = not bool(str(raw.get("Marked Deletion", "")).strip())
            obj.raw_data = raw

            if not existing:
                db.add(obj)
                inserted += 1
            else:
                updated += 1
        except Exception as e:
            errors.append({"row": row_num, "message": str(e)})

    db.commit()
    wb.close()
    return {"inserted": inserted, "updated": updated, "skipped": 0, "errors": errors}


def _upload_freight(wb, db: Session):
    """Parse multi-sheet freight Excel. Handles NAC Freight, Destination, and FAK sheets."""
    inserted, updated, errors = 0, 0, []
    skip_sheets = {"Cover Page", "Terms & Conditions Clauses", "Charge Clauses"}

    # Collect NAC and Destination data keyed by (origin, destination, container)
    nac_data = {}   # key -> raw_data dict
    dest_data = {}  # key -> raw_data dict
    fak_data = {}   # key -> raw_data dict

    for sn in wb.sheetnames:
        if sn.strip() in skip_sheets:
            continue
        ws = wb[sn]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 3:
            continue

        # Row 2 = headers (row 1 is group headers, row 3 is units)
        headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(rows[1])]
        is_nac = "NAC" in sn.upper() or "FREIGHT" in sn.upper()
        is_dest = "DESTINATION" in sn.upper()
        is_fak = "FAK" in sn.upper()

        # Track last seen general info for rows that inherit from above
        last_general = {}
        general_cols = ["Country of Loading", "Port of Loading", "Port of Destination", "Incoterm"]

        for row_num, row in enumerate(rows[3:], start=4):  # Data starts at row 4
            try:
                raw = {h: _cell_val(v) for h, v in zip(headers, row)}

                # Fill forward general info columns
                for gc in general_cols:
                    if gc in raw and raw[gc]:
                        last_general[gc] = raw[gc]
                    elif gc in last_general:
                        raw[gc] = last_general[gc]

                origin = str(raw.get("Port of Loading", "")).strip()
                dest = str(raw.get("Port of Destination", "")).strip()
                container = str(raw.get("Container", "")).strip()

                if not container:
                    continue

                raw["_sheet"] = sn
                key = (origin, dest, container)

                if is_dest:
                    dest_data[key] = raw
                elif is_fak:
                    fak_data[key] = raw
                else:
                    nac_data[key] = raw
            except Exception as e:
                errors.append({"row": row_num, "message": f"[{sn}] {e}"})

    # Merge NAC + Destination for same route, then process FAK separately
    all_keys = set(nac_data.keys()) | set(dest_data.keys()) | set(fak_data.keys())
    for key in all_keys:
        origin, dest, container = key
        merged = {}
        if key in nac_data:
            merged.update({"nac_charges": nac_data[key]})
            base = nac_data[key]
        if key in dest_data:
            merged.update({"dest_charges": dest_data[key]})
            base = dest_data.get(key, nac_data.get(key, {}))
        if key in fak_data:
            merged.update({"fak_charges": fak_data[key]})
            base = fak_data.get(key, {})
        if not merged:
            continue

        base = nac_data.get(key) or dest_data.get(key) or fak_data.get(key, {})

        existing = db.query(FreightRateCard).filter(
            FreightRateCard.origin == str(base.get("Country of Loading", "")).strip(),
            FreightRateCard.destination == dest,
            FreightRateCard.container_type == container,
        ).first()

        obj = existing or FreightRateCard(id=new_id("FRC-"))
        obj.origin = str(base.get("Country of Loading", "")).strip()
        obj.destination = dest
        obj.origin_port = origin
        obj.dest_port = dest
        obj.container_type = container
        obj.shipping_line = str(base.get("Shipping Line", "")).strip()
        obj.incoterm = str(base.get("Incoterm", "")).strip()
        obj.freight_currency = str(base.get("Freight Currency", "")).strip() or "USD"
        obj.dest_currency = str((dest_data.get(key) or {}).get("Destination Curerncy", "")).strip() or "AUD"
        obj.raw_data = merged

        # Set rate from ocean freight if available
        nac = nac_data.get(key) or fak_data.get(key, {})
        of_val = nac.get("Ocean Freight ", nac.get("Ocean Freight", 0))
        try:
            obj.rate = float(of_val) if of_val and str(of_val).strip() not in ("", "-", "Incl.", "n/a") else 0.0
        except (ValueError, TypeError):
            obj.rate = 0.0

        if not existing:
            db.add(obj)
            inserted += 1
        else:
            updated += 1

    db.commit()
    wb.close()
    return {"inserted": inserted, "updated": updated, "skipped": 0, "errors": errors}


def _upload_service(wb, db: Session):
    """Parse multi-sheet service rate card. Each sheet = one contractor. Skip 'Parts Time'."""
    inserted, updated, errors = 0, 0, []
    skip_sheets = {"Parts Time"}

    for sn in wb.sheetnames:
        sn_stripped = sn.strip()
        if sn_stripped in skip_sheets or sn_stripped.startswith("Sheet"):
            continue

        ws = wb[sn]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            continue

        headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(rows[0])]

        for row_num, row in enumerate(rows[1:], start=2):
            try:
                raw = {h: _cell_val(v) for h, v in zip(headers, row)}
                ft = str(raw.get("Fee Type", raw.get("Fee Type ", ""))).strip()
                if not ft:
                    continue

                raw["_contractor"] = sn_stripped
                val = raw.get("$Value", raw.get("$Value ", 0))
                rate_val = float(val) if val and val != "" else 0.0
                charge_desc = str(raw.get("Charge Rate", "")).strip()

                existing = db.query(ServiceRateCard).filter(
                    ServiceRateCard.contractor_name == sn_stripped,
                    ServiceRateCard.fee_type == ft,
                ).first()

                obj = existing or ServiceRateCard(id=new_id("SRC-"))
                obj.contractor_name = sn_stripped
                obj.fee_type = ft
                obj.service = ft
                obj.rate = rate_val
                obj.charge_description = charge_desc
                obj.raw_data = raw

                if not existing:
                    db.add(obj)
                    inserted += 1
                else:
                    updated += 1
            except Exception as e:
                errors.append({"row": row_num, "message": f"[{sn}] {e}"})

    db.commit()
    wb.close()
    return {"inserted": inserted, "updated": updated, "skipped": 0, "errors": errors}


def _upload_approval(wb, db: Session):
    """Parse approval sequence Excel. Columns: Category, Level 1 Approver, Level 2 Approver."""
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"inserted": 0, "updated": 0, "skipped": 0, "errors": []}

    headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(rows[0])]
    inserted, updated, errors = 0, 0, []

    for row_num, row in enumerate(rows[1:], start=2):
        try:
            raw = {h: _cell_val(v) for h, v in zip(headers, row)}
            category = str(raw.get("Category", "")).strip()
            if not category:
                continue

            l1 = str(raw.get("Level 1 Approver", "")).strip()
            l2 = str(raw.get("Level 2 Approver", "")).strip()

            existing = db.query(ApprovalSequenceMaster).filter(
                ApprovalSequenceMaster.name == category
            ).first()

            obj = existing or ApprovalSequenceMaster(id=new_id("ASM-"), name=category)
            obj.steps = [
                {"stepNumber": 1, "approverName": l1, "approverRole": "L1"},
                {"stepNumber": 2, "approverName": l2, "approverRole": "L2"},
            ]
            obj.raw_data = raw

            if not existing:
                db.add(obj)
                inserted += 1
            else:
                updated += 1
        except Exception as e:
            errors.append({"row": row_num, "message": str(e)})

    db.commit()
    wb.close()
    return {"inserted": inserted, "updated": updated, "skipped": 0, "errors": errors}


# ---------------------------------------------------------------------------
# Download — GET /api/masters/{master_type}/download
# ---------------------------------------------------------------------------
@router.get("/{master_type}/download")
def download_master(master_type: str, db: Session = Depends(get_db)):
    if master_type not in MASTER_TYPES:
        raise HTTPException(400, f"Invalid master_type: {master_type}. Must be one of: {list(MASTER_TYPES.keys())}")

    model_cls = MASTER_TYPES[master_type]
    rows = db.query(model_cls).all()

    wb = Workbook()
    ws = wb.active
    ws.title = master_type

    if not rows:
        ws.append(["No data"])
    elif rows[0].raw_data:
        # Reconstruct from raw_data — preserves original SAP columns
        if master_type == "freight-rate-cards":
            _download_freight(ws, rows)
        elif master_type == "service-rate-cards":
            _download_service(wb, rows)
        else:
            # Flat raw_data (vendors, approval)
            sample = rows[0].raw_data
            headers = [k for k in sample.keys() if not k.startswith("_")]
            ws.append(headers)
            for r in rows:
                rd = r.raw_data or {}
                ws.append([rd.get(h, "") for h in headers])
    else:
        # Legacy rows without raw_data — use to_dict
        sample = rows[0].to_dict()
        headers = list(sample.keys())
        ws.append(headers)
        for r in rows:
            d = r.to_dict()
            ws.append([d.get(h, "") for h in headers])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"{master_type}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _download_freight(ws, rows):
    """Reconstruct freight data — combine nac/dest/fak charges into columns."""
    # Collect all unique column names across all raw_data charge dicts
    all_cols = set()
    for r in rows:
        rd = r.raw_data or {}
        for charge_type in ("nac_charges", "dest_charges", "fak_charges"):
            if charge_type in rd:
                all_cols.update(k for k in rd[charge_type].keys() if not k.startswith("_"))
    cols = sorted(all_cols)
    ws.append(cols)
    for r in rows:
        rd = r.raw_data or {}
        # Merge all charge dicts, nac first then dest then fak
        merged = {}
        for ct in ("nac_charges", "dest_charges", "fak_charges"):
            if ct in rd:
                merged.update({k: v for k, v in rd[ct].items() if not k.startswith("_")})
        ws.append([merged.get(c, "") for c in cols])


def _download_service(wb, rows):
    """Reconstruct service rate card — one sheet per contractor."""
    ws = wb.active
    # Group rows by contractor
    by_contractor = {}
    for r in rows:
        cn = r.contractor_name or "Unknown"
        by_contractor.setdefault(cn, []).append(r)

    first = True
    for contractor, crows in by_contractor.items():
        if first:
            ws.title = contractor
            sheet = ws
            first = False
        else:
            sheet = wb.create_sheet(title=contractor[:31])  # Excel sheet name max 31 chars

        sample_rd = crows[0].raw_data or {}
        headers = [k for k in sample_rd.keys() if not k.startswith("_")]
        if not headers:
            headers = ["Fee Type", "Charge Rate", "$Value"]
        sheet.append(headers)
        for r in crows:
            rd = r.raw_data or {}
            if rd:
                sheet.append([rd.get(h, "") for h in headers])
            else:
                sheet.append([r.fee_type, r.charge_description, r.rate])
